#####################################################################################################################
#####################################################################################################################

" Internal Database Manager b1.0"

" 		Programmed by Tom McDonough, inspired by Kawika Tarayao, CFA 												"
"																													"
"																													"
"	USAGE																											"
"	**Used to automatically upload data from 100+ internal financial models into a central db. The program     **   "
"	**checks the database folder for existing excel files every 5 minutes, uploads applicable data from such   **   "
"	**files into a master file, and then removes the excel files from the folder. Prior to uploading, it will  **   "
"	**archive both the inbound file and the master file.                                                       **	"
"																													"
"	PURPOSE																											"
"	**The IDM should prove more versatile than simply linking disparate models through excel. The 'shotgun'    **	"
"   **functionality can be transfered to a newly made file simply by copying over the 'shotgun' worksheet into **	"
"	**any new file and then dropping the model into the database manager folder. It will also be more reliable,**	"
"	**since autoarchiving is built in, and less inking means less broken links, which in turn means more       **	"
"   **stability. Lastly, the autoarchiving function should lead to more reliable time series data for research **   "
"   **and analysis.                                 														   **   "

#####################################################################################################################
#####################################################################################################################

" Notes "

" Current status (7:40PM on 1/9/15): the script pulls data from any and all excel files in a given folder based on  "
" given cell coordinates. It then places that data into a master file per another set of coordinates. 		        "
" It does retain formulas in the master file. It also archives a copy of the master before beginning work, archives "
" a copy of the inbound file after it's done, and removes the inbound file from the applicable folder. It will 	    "
" repeat until all files are pulled from/removed from the folder, and will then wait five minutes before checking   "                                                
" the folder again.                                                                                 				"
"																													"
" Issues/places for improvement:																					"
"																													"
" #1 the formatting is not fully retained in the master file.  														"
" #2 I have not tested with .xls for older models																	"
" #3 Coordinates should be dynamic based on tags																	"
" 	#3a e.g., the program should look in the excel file for a keyword to determine what type of shotgun file it is, "
"     	and then look into a dictionary for the applicable cell coordinates 										"
" 	#3b e.g., the program should look in the excel file for the TICKER symbol and should be able to match that with "
"     	a dynamic list of destination coordinates. This must be dynamic (i.e., each time the program runs, it should"
"     	be able to locate each ticker's row/column) because tickers will often be added to the master file.         "
"																													"

#####################################################################################################################
#####################################################################################################################

" Modules "

import os
import time
import shutil
import sys
import xlrd
import xlwt
import datetime
from xlutils.copy import copy
from xlrd import *
from openpyxl import workbook
from openpyxl import load_workbook

#####################################################################################################################
#####################################################################################################################

" Global Variables "

masterFileAddress = 'c:\\Users\\MGHTSK\\My Documents\\Code\\Shotgun\\ShotgunTest_MasterFile\\na navs.xlsx' #
workingMasterFileAddress = 'c:\\Users\\MGHTSK\\My Documents\\Code\\Shotgun\\ShotgunTest_MasterFile\\na navs_working.xlsx' 
archiveDir = 'c:\\Users\\MGHTSK\\My Documents\\Code\\Shotgun\\ShotgunTest_Archive\\'
masterDir = 'c:\\Users\\MGHTSK\\My Documents\\Code\\Shotgun\\ShotgunTest_MasterFile\\'
shotgunDir = 'C:\\Users\\MGHTSK\\Documents\\Code\\Shotgun\\ShotgunTest\\'
extensionList = ['.xls', '.xlsx'] # this keeps track of which types of files we're looking for - unlikely to change
shotgunSheetName = 'shotgun'
masterSheetIndex = 0

" the 'desired data' items below may in the future become local, rather than global, variables                      "
" this would happen if, for example, i had a scraper that could read multiple different file types (e.g., NAV,      "
" Qs/Vs Database, FFO) based on a tag that was placed in the file                                                   "
"																													"
" THE LISTS SHOULD BE RESPECTIVE OF EACH OTHER (E.G., THE DATA FROM THE COORDINATES OF THE SECOND SET WITHIN THE    "
" DESIREDDATACOORDINATES LIST WILL BE PLACED IN THE COORDINATES OF THE SECOND SET WITHIN THE                        "
" destinationDataCoordinates LIST                                                                                   "

" the coordinates of the data desired from the original file (e.g., x,y) 											"
desiredDataCoordinates = [[2,5],[2,6],[2,7],[2,8],[2,9],[2,10],[2,11],[2,12],[2,13],[2,14],[2,15],[2,16],[2,17]] 

" the coordinates where the data will be placed in the new file (e.g., x,y) 										"
destinationDataCoordinates = [[2,13],[3,13],[4,13],[5,13],[6,13],[9,13],[10,13],[11,13],[12,13],[13,13],[14,13],[15,13],[16,13]]


#####################################################################################################################
#####################################################################################################################

" Functions "

def convertColumnNumToLetter(col):
	lettersInAlphabet = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
	col = int(col)
	if col <= 25:
		return lettersInAlphabet[col]
	else:
		colLetter1 = lettersInAlphabet[int(col/26)]
		colLetter2 = int(col%26)
		if colLetter2 != 0:
			colLetter2 = lettersInAlphabet[colLetter2]
		else:
			colLetter2 = lettersInAlphabet[colLetter2]
		return colLetter1+colLetter2

def convertColumnLetterToNum(col):
	lettersInAlphabet = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']
	if len(col) <= 1:
		return lettersInAlphabet.index(col)+1
	else:
		numberOfChars = len(col)
		columnNum = 0
		for i in range(0,numberOfChars):
			if i == 0:
				columnNum+=(numberOfChars-(i+1))*26*(lettersInAlphabet.index(col[i])+1)
				if i+1==numberOfChars:
					columnNum+=lettersInAlphabet.index(col[i])+1
			elif i == 1:
				columnNum+=(numberOfChars-(i+1))*26*(lettersInAlphabet.index(col[i])+1)
				if i+1==numberOfChars:
					columnNum+=lettersInAlphabet.index(col[i])+1
		return columnNum

def getExtension(excelFileLoc):
	return '.'+str(excelFileLoc.split('.')[1])

def getFileNameLessExt(excelFileLoc):
	return str(excelFileLoc.rsplit('.')[0].rsplit('\\')[1])

def getWorkbookRead(excelFile):
	extension = getExtension(excelFile)
	if extension == '.xls':
		return copy(open_workbook(excelFile, formatting_info=True))
	elif extension == '.xlsx':
		workbook = load_workbook(excelFile, use_iterators=False, data_only=True)
		return workbook

def getWorkbookWrite(excelFile):
	extension = getExtension(excelFile)
	if extension == '.xls':
		return copy(open_workbook(excelFile, formatting_info=True))
	elif extension == '.xlsx':
		workbook = load_workbook(excelFile)
		return workbook

def getSheetByName(workbook,extension,sheetName):
	if extension == '.xls':
		worksheet = workbook.sheet_by_name(sheetName)
		return worksheet
	elif extension == '.xlsx':
		worksheet = workbook.get_sheet_by_name(sheetName)
		return worksheet

def getSheetByIndex(workbook,extension,sheetIndex):
	if extension == '.xls':
		return workbook.sheet_by_index(sheetIndex)
	elif extension == '.xlsx':
		return workbook.active

def writeCell(worksheet,extension,row,col,data):
	if extension == '.xls':
		worksheet.write(row,col,data)
	elif extension == '.xlsx':
		col = convertColumnNumToLetter(col)
		excelCoord = str(col)+str(row+1)
		worksheet[excelCoord] = data

def numRowsInSheet(worksheet,extension):
	if extension == '.xls':
		return worksheet.nrows
	elif extension == '.xlsx':
		return len(worksheet.rows)

def numColsInSheet(worksheet,extension):
	if extension == '.xls':
		return worksheet.ncols
	elif extension == '.xlsx':
		return len(worksheet.rows[0])
		
def copyFileToArchive(excelFileLoc, archiveDir):

	"copys file that was just added to the master list into the archive directory and then deletes the file from the shotgun directory"

	extension = getExtension(excelFileLoc)
	
	archiveFileName = excelFileLoc[excelFileLoc.rfind('\\')+1:excelFileLoc.find('.')]+"_archive_"+str(datetime.date.today())+'_'+str(time.strftime('%H-%M-%S'))+extension
	shutil.copyfile(excelFileLoc,os.path.join(archiveDir,archiveFileName))
	os.remove(excelFileLoc)

def getListOfExcelFiles(directory):

	"will return a list of all excel files in the shotgun directory"

	excelFiles = [os.path.normcase(f) for f in os.listdir(directory)]
	excelFiles = [os.path.join(directory, f) for f in excelFiles if os.path.splitext(f)[1] in extensionList]
	return excelFiles

def getCurrentTimeInHours():
	
	"converts current time into a single numeric float in terms of hours that can be evaluated vs. desired time frame"
	
	timeInHours = float(time.strftime("%H"))
	timeInHours += float(time.strftime("%M"))/60
	timeInHours += float(time.strftime("%S"))/60/60
	return timeInHours

def copyMasterFileToArchive(masterDir, masterFileAddress, workingMasterFileAddress, archiveDir):
	
	"before initiating dump, this will make an archived copy of the previous master file"
	"after copying the master file over, it renames it until there are no remaining files in the shotgun folder,"
	"so that the program doesn't make a new archive of the master file for each file being shotgunned"

	filesInMasterDirectory = getListOfExcelFiles(masterDir)
	
	if masterFileAddress in filesInMasterDirectory:
		extension = masterFileExt
		archiveMasterFileName = masterFileNameLessExt+str(datetime.date.today())+'_'+str(time.strftime('%H-%M-%S'))+extension
		shutil.copyfile(masterFileAddress,os.path.join(archiveDir,archiveMasterFileName))
		os.rename(masterFileAddress,workingMasterFileAddress)

def renameMasterFile(masterDir, masterFileAddress, workingMasterFileAddress):
	" at the end of the loop, this will rename the old master file back from 'working' to its original name "
	filesInMasterDirectory = getListOfExcelFiles(masterDir)
	if workingMasterFileAddress in filesInMasterDirectory:
		os.rename(workingMasterFileAddress,masterFileAddress)

def moveDataToMaster(excelFileLoc):	

	" loads the worksheet using xlrd; copys the applicable data into a list; exports that data into the masterFile "

	excelFileExtension = getExtension(excelFileLoc)

	def writeDataToMaster(excelData, workingMasterFileAddress, desiredDataCoordinates, destinationDataCoordinates):

		" using the excel data in lists, writes specific data to another file"

		def pullDesiredData(excelData):
			pulledData = []
			for a in desiredDataCoordinates:
				pulledData.append(excelData[a[1]][a[0]-1]) # because the coordinates above are column, row, we first pull the row number to correspond with the correct list within the excel data list, and we then pull the column number to gather the right data point from within that row
			return pulledData

		def pushDesiredData(pulledData):
			masterFileData = getWorkbookWrite(workingMasterFileAddress)
			sheet = getSheetByIndex(masterFileData,masterFileExt,masterSheetIndex)
			for a in destinationDataCoordinates:
				row, col, data = a[1], a[0], pulledData[destinationDataCoordinates.index(a)]
				writeCell(sheet,masterFileExt,row,col,data)
			masterFileData.save(workingMasterFileAddress)

		pushDesiredData(pullDesiredData(excelData))

	def loadExcelData(excelFileLoc):

		def importSheet1(worksheet, numRows, numCols):
			
			def createMatrix(numRows):
				" Creates a list of lists, where the length of the parent list is the number of rows. The columns are to be filled in later"
				print numRows
				return [[] for a in range(0,numRows)]

			def fillMatrix(worksheet, extension, matrix, numRows, numCols):
				" Fills the list of list with the contents of the excel file "
				for rowNum in range(0,numRows):
					for colNum in range(0,numCols):
						if extension == '.xls':
							cell = worksheet.cell(rowx=rowNum,colx=colNum).value
						elif extension == '.xlsx':
							colLetter = convertColumnNumToLetter(colNum)
							cell = worksheet[colLetter+str(rowNum+1)].value
						matrix[rowNum].append(cell)
				print matrix
				return matrix
			# def cleanMatrix(matrix):

			# 	for row in matrix:
			# 		rowLoc = matrix.index(row)
			# 		for cell in row:
			# 			colLoc = row.index(cell)
			# 			if "u'" in row:
			# 				matrix[rowLoc][colLoc] = cell.split("u'")[]
			# 	return matrix

			storedData = createMatrix(numRows)
			storedData = fillMatrix(worksheet, excelFileExtension, storedData, numRows, numCols)
			# storedData = cleanMatrix(storedData)
			# print "cleaned matrix: ",storedData
			
			return storedData

		workbook = getWorkbookRead(excelFileLoc)
		
		worksheet = getSheetByName(workbook,excelFileExtension,shotgunSheetName)
		
		try:
			numRows = numRowsInSheet(worksheet,excelFileExtension)
		except:
			print "Could not find data to upload. Is your shotgun tab named correctly? (it should be titled '"+shotgunSheetName+"')"
			sys.exit()

		numCols = numColsInSheet(worksheet,excelFileExtension)

		return importSheet1(worksheet,numRows,numCols)

	inputData = loadExcelData(excelFileLoc)
	writeDataToMaster(inputData,workingMasterFileAddress,desiredDataCoordinates,destinationDataCoordinates)
	
#####################################################################################################################
#####################################################################################################################

" Declared Variables "

masterFileExt = getExtension(masterFileAddress)
masterFileNameLessExt = getFileNameLessExt(masterFileAddress)

#####################################################################################################################
#####################################################################################################################

" Script "

while getCurrentTimeInHours() > 7 and getCurrentTimeInHours() < 20:
	print "Step 1"
	while getListOfExcelFiles(shotgunDir): # runs until list of excel files is empty (i.e., no more excel files in folder) 
		print "Step 2"
		copyMasterFileToArchive(masterDir, masterFileAddress, workingMasterFileAddress, archiveDir)
		print "Step 3"
		fileToBeUploaded = getListOfExcelFiles(shotgunDir)[0]
		print "Step 4"
		moveDataToMaster(fileToBeUploaded)
		print "Step 5"
		copyFileToArchive(fileToBeUploaded, archiveDir)
		print "Step 6"
	
	renameMasterFile(masterDir, masterFileAddress, workingMasterFileAddress) # does not rename master file until no files remaining to update, so that duplicate archives of master file are not created
	time.sleep(60*5)


